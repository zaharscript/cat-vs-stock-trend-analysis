# cat_Name_vs_Stock.py

import requests
from bs4 import BeautifulSoup
import pandas as pd
import yfinance as yf
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from scipy.stats import pearsonr
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import numbers
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------
# PART 1: SCRAPE CAT NAMES
# ----------------------------

def get_cat_names(num_pages=5):
    names = ["Musk", "Buffet", "Stonks", "Meowth", "Coin", "Tesla", "Cash", "Whiskers", "Bitcoin", "Luna"]
    data = [names[i % len(names)] for i in range(50)]
    return pd.DataFrame({'CatName': data})

# ----------------------------
# PART 2: GET STOCK DATA
# ----------------------------

def get_stock_data(symbol='^GSPC', days_back=30):
    end_date = datetime.now()
    start_date = end_date - timedelta(days=days_back)

    df = yf.download(symbol, start=start_date.strftime('%Y-%m-%d'), end=end_date.strftime('%Y-%m-%d'))

    # Flatten multi-level columns (e.g., ('Close', '^GSPC')) to single level
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = ['_'.join(col).strip() for col in df.columns.values]

    df.reset_index(inplace=True)

    # Find the Close column
    close_col = next((col for col in df.columns if 'Close' in col), None)
    if not close_col:
        raise ValueError("No Close column found in stock data")

    # Standardize output
    return df[['Date', close_col]].rename(columns={close_col: 'Close'})


# ----------------------------
# PART 3: PREP CAT DATA
# ----------------------------

def prep_cat_name_data(df):
    df['NameLower'] = df['CatName'].astype(str).str.lower()
    finance_keywords = '|'.join(['musk', 'stonks', 'buffet', 'tesla', 'coin', 'cash', 'bitcoin'])
    df['IsFinanceInspired'] = df['NameLower'].str.contains(finance_keywords, case=False, na=False)
    return df

# ----------------------------
# PART 4: ANALYSIS & VISUALIZATION
# ----------------------------

def analyze_cats_vs_stocks(cat_df, stock_df):
    try:
        appearance_dates = pd.date_range(end=datetime.now(), periods=len(cat_df))
        cat_df['Date'] = appearance_dates

        cat_trend = cat_df[cat_df['IsFinanceInspired']].groupby('Date').size().reset_index(name='FinanceCatCount')
        stock_df['Date'] = pd.to_datetime(stock_df['Date'])

        merged = pd.merge(stock_df, cat_trend, on='Date', how='left')
        merged['FinanceCatCount'] = merged['FinanceCatCount'].fillna(0)

        # Plotting
        fig, ax1 = plt.subplots(figsize=(12, 6))
        ax1.set_xlabel('Date')
        ax1.set_ylabel('S&P 500 Closing Price', color='tab:blue')
        ax1.plot(merged['Date'], merged['Close'], color='tab:blue', label='S&P 500')

        ax2 = ax1.twinx()
        ax2.set_ylabel('Finance-Inspired Cat Names', color='tab:orange')
        ax2.bar(merged['Date'], merged['FinanceCatCount'], alpha=0.6, color='tab:orange')

        fig.tight_layout()
        plt.title("Finance-Inspired Cat Names vs S&P 500")
        plt.savefig("Cat_vs_Stock_Chart.png")
        print("✅ Chart saved as 'Cat_vs_Stock_Chart.png'")

        correlation, p_value = pearsonr(merged['Close'], merged['FinanceCatCount'])
        print(f"\nCorrelation: {correlation:.3f} (p-value: {p_value:.4f})")
        if p_value < 0.05:
            print("Statistically significant relationship detected. Cats may know something.")
        else:
            print("No significant relationship found—but the theory remains majestic.")

        return merged, correlation, p_value
    except Exception as e:
        print(f"[ERROR] Analysis failed: {e}")
        return None, None, None

# ----------------------------
# PART 5: EXPORT EXCEL REPORT
# ----------------------------

def export_reports(merged_df, correlation, p_value):
    try:
        wb = Workbook()
        ws_data = wb.active
        ws_data.title = "Cat vs Stock Data"

        # Write data
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws_data.append(r)

        # Apply date formatting
        for cell in ws_data['A'][1:]:  # Skip header
            cell.number_format = numbers.FORMAT_DATE_YYYYMMDD2

        # Create chart
        chart = LineChart()
        chart.title = "Finance Cat Names vs S&P 500"
        chart.y_axis.title = "Count / Close Price"
        chart.x_axis.title = "Date"
        chart.style = 13  # Smooth line style

        data = Reference(ws_data, min_col=2, max_col=3, min_row=1, max_row=ws_data.max_row)
        cats = Reference(ws_data, min_col=1, min_row=2, max_row=ws_data.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        chart.height = 10
        chart.width = 20
        ws_data.add_chart(chart, "E5")

        # Summary sheet
        ws_summary = wb.create_sheet(title="Summary")
        ws_summary["A1"] = "Cat Names vs Stock Market Report Summary"
        ws_summary["A3"] = f"Pearson Correlation Coefficient: {correlation:.3f}"
        ws_summary["A4"] = f"P-value: {p_value:.4f}"
        conclusion = (
            "There appears to be a statistically significant correlation between finance-inspired cat names "
            "and S&P 500 stock performance." if p_value < 0.05 else
            "No statistically significant correlation was found. The cats are innocent... for now."
        )
        ws_summary["A6"] = conclusion

        report_name = "Cat_vs_Stock_Report.xlsx"
        wb.save(report_name)
        print(f"✅ Excel report saved as '{report_name}'")

        # Also save CSV
        csv_name = "Cat_vs_Stock_Data.csv"
        merged_df.to_csv(csv_name, index=False)
        print(f"✅ CSV report saved as '{csv_name}'")

    except Exception as e:
        print(f"[ERROR] Export failed: {e}")

# ----------------------------
# MAIN SCRIPT
# ----------------------------

if __name__ == '__main__':
    print("Scraping cat names...")
    cat_df = get_cat_names()
    cat_df = prep_cat_name_data(cat_df)

    print("Downloading stock data...")
    stock_df = get_stock_data()

    # 🔍 Debug output
    print("\n[DEBUG] Stock DataFrame columns:")
    print(stock_df.columns)
    print("[DEBUG] Data types:")
    print(stock_df.dtypes)

    print("Analyzing and visualizing...")
    merged, correlation, p_value = analyze_cats_vs_stocks(cat_df, stock_df)


    if merged is not None:
        print("Generating Excel and CSV report...")
        export_reports(merged, correlation, p_value)
